#!/usr/bin/env python3
"""
Bulk Neuron north MQTT subscriptions (and optional tags) from an Excel template.

Topic pattern (when topic_override is empty):
  factory/{site}/{area}/{line}/{leaf}/{data_type}
  - leaf = slug(device_id) if device_id non-empty, else slug(group)
  - Empty site defaults to quang-ninh; empty data_type defaults to telemetry
  - area and line are omitted from the path when their cells are empty

MQTT north plugin uses one QoS for telemetry publishes; retain per data_type is not
supported in-plugin — use EMQX rules or a second north app if needed.

Dependencies: pip install -r requirements.txt (openpyxl)
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.error
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("Missing openpyxl. Run: pip install -r requirements.txt", file=sys.stderr)
    sys.exit(1)

DEFAULT_SITE = "quang-ninh"
DEFAULT_DATA_TYPE = "telemetry"
INVALID_TOPIC_CHARS = re.compile(r"[#+]")


def slug(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^a-z0-9._-]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s


def normalize_header(cell: Any) -> str:
    if cell is None:
        return ""
    t = str(cell).strip().lower().replace(" ", "_")
    aliases = {
        "node": "driver",
        "south_node": "driver",
        "tag": "tag_name",
        "tagname": "tag_name",
        "name": "tag_name",
        "interval": "interval_ms",
    }
    return aliases.get(t, t)


def read_header_map(ws: Worksheet, row: int = 1) -> Dict[str, int]:
    """Map normalized column name -> 1-based column index."""
    m: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[row], start=1):
        key = normalize_header(cell.value)
        if key:
            m[key] = col_idx
    return m


def cell_str(ws: Worksheet, row: int, col_map: Dict[str, int], key: str) -> str:
    if key not in col_map:
        return ""
    v = ws.cell(row=row, column=col_map[key]).value
    if v is None:
        return ""
    return str(v).strip()


def cell_optional_int(ws: Worksheet, row: int, col_map: Dict[str, int], key: str) -> Optional[int]:
    s = cell_str(ws, row, col_map, key)
    if s == "":
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def cell_optional_float(ws: Worksheet, row: int, col_map: Dict[str, int], key: str) -> Optional[float]:
    s = cell_str(ws, row, col_map, key)
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def build_topic(
    *,
    site: str,
    area: str,
    line: str,
    leaf: str,
    data_type: str,
    topic_override: str,
) -> str:
    if topic_override:
        return topic_override.strip()
    site_f = slug(site) if site else slug(DEFAULT_SITE)
    parts = ["factory", site_f]
    if area:
        parts.append(slug(area))
    if line:
        parts.append(slug(line))
    parts.append(slug(leaf))
    dt = slug(data_type) if data_type else DEFAULT_DATA_TYPE
    if not dt:
        dt = DEFAULT_DATA_TYPE
    parts.append(dt)
    topic = "/".join(parts)
    if INVALID_TOPIC_CHARS.search(topic):
        raise ValueError(f"Topic contains invalid MQTT wildcard chars: {topic!r}")
    return topic


@dataclass
class MappingRow:
    row_num: int
    driver: str
    group: str
    site: str
    area: str
    line: str
    device_id: str
    data_type: str
    topic_override: str

    def resolved_topic(self) -> str:
        leaf_src = self.device_id.strip() if self.device_id.strip() else self.group
        site = self.site.strip() if self.site.strip() else DEFAULT_SITE
        dt = self.data_type.strip() if self.data_type.strip() else DEFAULT_DATA_TYPE
        return build_topic(
            site=site,
            area=self.area,
            line=self.line,
            leaf=leaf_src,
            data_type=dt,
            topic_override=self.topic_override,
        )


def load_mapping_rows(ws: Worksheet) -> Tuple[List[MappingRow], List[str]]:
    errors: List[str] = []
    if ws.max_row < 2:
        return [], ["Mapping sheet has no data rows"]
    col_map = read_header_map(ws, 1)
    required = {"driver", "group"}
    missing = required - set(col_map.keys())
    if missing:
        return [], [f"Mapping sheet missing columns: {sorted(missing)}"]

    rows: List[MappingRow] = []
    seen: Dict[Tuple[str, str], int] = {}
    for r in range(2, ws.max_row + 1):
        if all(
            ws.cell(row=r, column=col_map.get(k, 0)).value in (None, "")
            for k in ("driver", "group")
            if k in col_map
        ):
            # skip completely empty rows
            if not any(ws.cell(row=r, column=c).value not in (None, "") for c in col_map.values()):
                continue
        driver = cell_str(ws, r, col_map, "driver")
        group = cell_str(ws, r, col_map, "group")
        if not driver or not group:
            errors.append(f"Row {r}: driver and group are required")
            continue
        key = (driver, group)
        if key in seen:
            errors.append(
                f"Row {r}: duplicate (driver={driver!r}, group={group!r}); "
                f"keeping row {seen[key]}, skipping"
            )
            continue
        seen[key] = r
        rows.append(
            MappingRow(
                row_num=r,
                driver=driver,
                group=group,
                site=cell_str(ws, r, col_map, "site"),
                area=cell_str(ws, r, col_map, "area"),
                line=cell_str(ws, r, col_map, "line"),
                device_id=cell_str(ws, r, col_map, "device_id"),
                data_type=cell_str(ws, r, col_map, "data_type"),
                topic_override=cell_str(ws, r, col_map, "topic_override"),
            )
        )
    return rows, errors


@dataclass
class TagRow:
    row_num: int
    driver: str
    group: str
    interval_ms: Optional[int]
    name: str
    address: str
    attribute: int
    type: int
    precision: Optional[int]
    decimal: Optional[float]
    description: str


def load_tag_rows(ws: Worksheet) -> Tuple[List[TagRow], List[str]]:
    errors: List[str] = []
    if ws.max_row < 2:
        return [], []
    col_map = read_header_map(ws, 1)
    need = {"driver", "group", "tag_name", "address", "attribute", "type"}
    if not need.issubset(col_map.keys()):
        return [], [
            f"Tags sheet missing columns (need {sorted(need)}), have {sorted(col_map.keys())}"
        ]

    rows: List[TagRow] = []
    for r in range(2, ws.max_row + 1):
        driver = cell_str(ws, r, col_map, "driver")
        group = cell_str(ws, r, col_map, "group")
        name = cell_str(ws, r, col_map, "tag_name")
        address = cell_str(ws, r, col_map, "address")
        if not driver and not group and not name and not address:
            continue
        if not driver or not group or not name or not address:
            errors.append(f"Tags row {r}: driver, group, tag_name, address are required")
            continue
        attr_s = cell_str(ws, r, col_map, "attribute")
        type_s = cell_str(ws, r, col_map, "type")
        try:
            attribute = int(float(attr_s))
            typ = int(float(type_s))
        except ValueError:
            errors.append(f"Tags row {r}: attribute and type must be integers")
            continue
        rows.append(
            TagRow(
                row_num=r,
                driver=driver,
                group=group,
                interval_ms=cell_optional_int(ws, r, col_map, "interval_ms"),
                name=name,
                address=address,
                attribute=attribute,
                type=typ,
                precision=cell_optional_int(ws, r, col_map, "precision"),
                decimal=cell_optional_float(ws, r, col_map, "decimal"),
                description=cell_str(ws, r, col_map, "description"),
            )
        )
    return rows, errors


def group_tags_by_node(tag_rows: List[TagRow]) -> Dict[str, List[Dict[str, Any]]]:
    """Merge tags per (driver, group), then group payloads by south node name for /api/v2/gtags."""
    by_key: Dict[Tuple[str, str], Dict[str, Any]] = {}
    intervals: Dict[Tuple[str, str], int] = defaultdict(lambda: 3000)

    for tr in tag_rows:
        key = (tr.driver, tr.group)
        if key not in by_key:
            by_key[key] = {"group": tr.group, "interval": 3000, "tags": []}
        if tr.interval_ms is not None:
            intervals[key] = max(intervals[key], tr.interval_ms)
        tag_obj: Dict[str, Any] = {
            "name": tr.name,
            "address": tr.address,
            "attribute": tr.attribute,
            "type": tr.type,
        }
        if tr.precision is not None:
            tag_obj["precision"] = tr.precision
        if tr.decimal is not None:
            tag_obj["decimal"] = tr.decimal
        if tr.description:
            tag_obj["description"] = tr.description
        by_key[key]["tags"].append(tag_obj)

    by_node: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for (node, _g), g in by_key.items():
        g["interval"] = intervals[(node, _g)]
        by_node[node].append(g)
    return by_node


def http_post_json(
    url: str, token: str, body: Dict[str, Any], timeout: float = 60.0
) -> Tuple[int, Dict[str, Any]]:
    data = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {token}",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
            code = resp.getcode()
    except urllib.error.HTTPError as e:
        raw = e.read().decode("utf-8", errors="replace")
        code = e.code
    try:
        parsed: Dict[str, Any] = json.loads(raw) if raw.strip() else {}
    except json.JSONDecodeError:
        parsed = {"_raw": raw}
    return code, parsed


def cmd_generate_template(path: str) -> None:
    wb = Workbook()
    # Mapping
    m = wb.active
    m.title = "Mapping"
    headers = [
        "driver",
        "group",
        "site",
        "area",
        "line",
        "device_id",
        "data_type",
        "topic_override",
    ]
    for i, h in enumerate(headers, start=1):
        m.cell(row=1, column=i, value=h)
    # Sample row: driver TC2_2, group BNS12 -> factory/quang-ninh/hoanhbo/tc2-2/bns12/telemetry
    # (leaf from slug(group); device_id empty)
    example = [
        "TC2_2",
        "BNS12",
        "quang-ninh",
        "hoanhbo",
        "tc2-2",
        "",
        "telemetry",
        "",
    ]
    for i, v in enumerate(example, start=1):
        m.cell(row=2, column=i, value=v)
    m.freeze_panes = "A2"

    # Tags
    t = wb.create_sheet("Tags")
    theaders = [
        "driver",
        "group",
        "interval_ms",
        "tag_name",
        "address",
        "attribute",
        "type",
        "precision",
        "decimal",
        "description",
    ]
    for i, h in enumerate(theaders, start=1):
        t.cell(row=1, column=i, value=h)
    tex = [
        "TC2_2",
        "BNS12",
        "3000",
        "example_tag",
        "1!40001",
        "3",
        "9",
        "2",
        "",
        "Example — edit address/types for your PLC",
    ]
    for i, v in enumerate(tex, start=1):
        t.cell(row=2, column=i, value=v)
    t.freeze_panes = "A2"

    wb.save(path)
    print(f"Wrote template: {path}")


def cmd_import(args: argparse.Namespace) -> int:
    path = args.excel
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except OSError as e:
        print(f"Cannot open {path}: {e}", file=sys.stderr)
        return 1

    if "Mapping" not in wb.sheetnames:
        print(f"No 'Mapping' sheet in {path}. Sheets: {wb.sheetnames}", file=sys.stderr)
        return 1

    wb.close()
    wb = load_workbook(path, data_only=True)
    mapping_ws = wb["Mapping"]
    map_rows, map_err = load_mapping_rows(mapping_ws)
    if map_err and not map_rows:
        for e in map_err:
            print(e, file=sys.stderr)
        return 1
    for e in map_err:
        print(f"Warning: {e}", file=sys.stderr)

    tags_by_node: Dict[str, List[Dict[str, Any]]] = {}
    if "Tags" in wb.sheetnames and not args.skip_tags:
        tag_rows, tag_err = load_tag_rows(wb["Tags"])
        for e in tag_err:
            print(f"Tags error: {e}", file=sys.stderr)
        if tag_err:
            wb.close()
            return 1
        tags_by_node = group_tags_by_node(tag_rows)
    wb.close()

    base = args.base_url.rstrip("/")
    token = args.token
    app = args.app
    batch = max(1, int(args.batch_size))

    # Preview topics
    for mr in map_rows[:5]:
        print(f"  row {mr.row_num}: {mr.driver}/{mr.group} -> {mr.resolved_topic()}")
    if len(map_rows) > 5:
        print(f"  ... and {len(map_rows) - 5} more mapping rows")

    if args.dry_run:
        print("Dry run: no HTTP requests sent.")
        if tags_by_node:
            nreq = sum(
                (len(gl) + batch - 1) // batch for gl in tags_by_node.values()
            )
            print(f"Would send ~{nreq} gtags HTTP request(s) (batch-size={batch}).")
        return 0

    # POST gtags first (optional): chunk groups[] per node
    if tags_by_node:
        for node, groups in tags_by_node.items():
            for i in range(0, len(groups), batch):
                chunk = groups[i : i + batch]
                body = {"node": node, "groups": chunk}
                url = f"{base}/api/v2/gtags"
                code, resp = http_post_json(url, token, body)
                err = resp.get("error", -1)
                if code != 200 or err != 0:
                    print(
                        f"gtags POST failed HTTP {code} error={err} body={resp}",
                        file=sys.stderr,
                    )
                    return 1
                print(f"gtags OK node={node!r} groups={len(chunk)}")

    # POST subscribes in batches
    url = f"{base}/api/v2/subscribes"
    for i in range(0, len(map_rows), batch):
        chunk = map_rows[i : i + batch]
        groups = []
        for mr in chunk:
            groups.append(
                {
                    "driver": mr.driver,
                    "group": mr.group,
                    "params": {"topic": mr.resolved_topic()},
                }
            )
        body = {"app": app, "groups": groups}
        code, resp = http_post_json(url, token, body)
        err = resp.get("error", -1)
        if code != 200 or err != 0:
            print(f"subscribes POST failed HTTP {code} error={err} body={resp}", file=sys.stderr)
            return 1
        print(f"subscribes OK rows {i + 1}-{i + len(chunk)} / {len(map_rows)}")

    return 0


def main() -> int:
    p = argparse.ArgumentParser(description="Neuron MQTT mapping Excel import")
    sub = p.add_subparsers(dest="cmd", required=True)

    g = sub.add_parser("generate-template", help="Write a sample .xlsx template")
    g.add_argument("-o", "--output", required=True, help="Output .xlsx path")

    i = sub.add_parser("import", help="Read Excel and POST to Neuron API")
    i.add_argument("--base-url", required=True, help="Neuron API base, e.g. http://127.0.0.1:7001")
    i.add_argument("--token", required=True, help="Bearer token for Authorization")
    i.add_argument("--app", default="mqtt", help="North MQTT app/node name (default: mqtt)")
    i.add_argument("--excel", required=True, help="Path to .xlsx file")
    i.add_argument("--batch-size", type=int, default=80, help="Groups per HTTP request (default: 80)")
    i.add_argument("--dry-run", action="store_true", help="Parse only; do not call API")
    i.add_argument(
        "--skip-tags",
        action="store_true",
        help="Ignore Tags sheet even if present",
    )

    args = p.parse_args()
    if args.cmd == "generate-template":
        cmd_generate_template(args.output)
        return 0
    return int(cmd_import(args))


if __name__ == "__main__":
    sys.exit(main())
